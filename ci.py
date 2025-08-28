import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO
import zipfile

# ---------- Utils ----------
def safe_float(x):
    if x is None:
        return 0.0
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except:
        return 0.0

def load_data_from_gsheet():
    # Sheet Data cột lần lượt:
    # 0 SKU | 1 Status_steel | 2 Weight_steel | 3 Cost_steel
    # 4 Status_alu | 5 Weight_alu | 6 Cost_alu
    url = "https://docs.google.com/spreadsheets/d/1sLEMyyTbjumWn5TpYV6cZKxsUO5l10umZSqIYByU_Bk/export?format=csv&gid=2043981257"
    df = pd.read_csv(url, dtype=str).fillna("")
    data_map = {}
    for _, row in df.iterrows():
        sku = row.iloc[0].strip()
        if not sku:
            continue
        data_map[sku] = {
            "steelStatus": row.iloc[1].strip(),
            "steelWeight": safe_float(row.iloc[2]),
            "steelCost":   safe_float(row.iloc[3]),
            "aluStatus":   row.iloc[4].strip(),
            "aluWeight":   safe_float(row.iloc[5]),
            "aluCost":     safe_float(row.iloc[6]),
        }
    return data_map

# ---------- Core ----------
def process_excel(file, data_map):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.worksheets[0]   # lấy sheet đầu tiên

    # Cấu trúc vùng dữ liệu như trước
    DATA_HEADER_ROW = 10
    DATA_START = 11
    FOOTER_ROWS = 5
    last_row = ws.max_row
    footer_start = last_row - FOOTER_ROWS + 1   # TOTAL nằm trong đoạn này

    # Lấy data trước TOTAL (file gốc đang có 7 cột)
    data_rows = list(ws.iter_rows(min_row=DATA_START,
                                  max_row=footer_start-2,
                                  min_col=1, max_col=7,
                                  values_only=True))

    new_rows = []
    merge_groups = []
    cur_row = DATA_START

    for row in data_rows:
        if not row:
            continue
        po, sku, asin, desc, qty, price, _ = (list(row) + [None]*7)[:7]
        sku  = (sku or "").strip() if isinstance(sku, str) else str(sku or "").strip()
        desc = "" if desc is None else str(desc)
        qty_f   = safe_float(qty)
        price_f = safe_float(price)

        # Bỏ qua dòng rỗng
        if not sku and qty_f == 0 and price_f == 0 and not desc:
            continue

        # SKU không có trong Data → giữ nguyên, thêm 2 cột weight rỗng
        if not sku or sku not in data_map:
            new_rows.append([po, sku, asin, desc, qty_f, price_f, qty_f*price_f, None, None])
            cur_row += 1
            continue

        m = data_map[sku]
        steel_on = m["steelStatus"].lower() == "yes"
        alu_on   = m["aluStatus"].lower() == "yes"
        steel_cost   = m["steelCost"] if steel_on else 0.0
        alu_cost     = m["aluCost"]   if alu_on   else 0.0
        steel_weight = m["steelWeight"] if steel_on else 0.0
        alu_weight   = m["aluWeight"]   if alu_on   else 0.0

        group_start = cur_row

        # Dòng gốc (giá đã trừ cost steel/alu), weight để trống
        new_price  = price_f - steel_cost - alu_cost
        new_rows.append([po, sku, asin, desc, qty_f, new_price, qty_f*new_price, None, None])
        cur_row += 1

        # Dòng steel
        if steel_on:
            new_rows.append([po, sku, asin, f"{desc}_steel", qty_f, steel_cost, qty_f*steel_cost, steel_weight, None])
            cur_row += 1

        # Dòng aluminum
        if alu_on:
            new_rows.append([po, sku, asin, f"{desc}_aluminum", qty_f, alu_cost, qty_f*alu_cost, None, alu_weight])
            cur_row += 1

        if cur_row - 1 > group_start:
            merge_groups.append((group_start, cur_row - 1))

    # Tính chênh lệch số dòng & chèn/xoá trước khi ghi
    old_data_len = (footer_start - 2) - DATA_START + 1
    new_data_len = len(new_rows)
    delta = new_data_len - old_data_len
    if delta > 0:
        ws.insert_rows(footer_start-1, amount=delta)
        footer_start += delta
    elif delta < 0:
        ws.delete_rows(idx=DATA_START + new_data_len, amount=-delta)
        footer_start += delta   # delta < 0 nên dịch ngược lên

    total_row = footer_start - 1

    # Ghi header CHUẨN 9 cột
    HEADERS = [
        "PO", "SKU", "ASIN", "DESCRIPTION",
        "QUANTITY", "UNIT PRICE (USD)", "AMOUNT",
        "weight_steel", "weight_alu"
    ]
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(DATA_HEADER_ROW, c, value=name)
        cell.font = Font(bold=True)

    # Ghi data mới (9 cột)
    for i, r in enumerate(new_rows):
        for c, v in enumerate(r, start=1):
            ws.cell(row=DATA_START + i, column=c, value=v)

    # Bỏ merge cũ trong vùng data
    to_unmerge = [rng for rng in ws.merged_cells.ranges
                  if not (rng.max_row < DATA_START or rng.min_row >= footer_start)]
    for rng in to_unmerge:
        ws.unmerge_cells(str(rng))

    # Merge lại các cột 1,2,3,5 theo nhóm (PO, SKU, ASIN, QUANTITY)
    for start, end in merge_groups:
        for col in (1, 2, 3, 5):
            ws.merge_cells(start_row=start, end_row=end, start_column=col, end_column=col)

    # Border + highlight (bao trọn header -> TOTAL), áp dụng cho 9 cột
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for r in range(DATA_HEADER_ROW, total_row + 1):
        for c in range(1, 10):   # 1..9
            cell = ws.cell(r, c)
            cell.border = border
            # Tô vàng dòng có Unit Price = 0 (trừ header và TOTAL)
            if r > DATA_HEADER_ROW and r < total_row and safe_float(ws.cell(r, 6).value) == 0:
                cell.fill = fill_yellow

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ---------- Streamlit UI ----------
st.title("Tool check CI")
uploaded = st.file_uploader("Upload Excel (.xlsx)", type="xlsx", accept_multiple_files=True)
data_map = load_data_from_gsheet()

# nút Run
if st.button("▶️ Run") and uploaded:
    if len(uploaded) == 1:
        f = uploaded[0]
        out = process_excel(f, data_map)
        fname = f.name.replace(".xlsx", "_checked.xlsx")
        st.download_button(
            label="⬇️ Download Excel",
            data=out,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for f in uploaded:
                out = process_excel(f, data_map)
                fname = f.name.replace(".xlsx", "_checked.xlsx")
                zipf.writestr(fname, out.getvalue())
        zip_buffer.seek(0)

        st.download_button(
            label="📦 Download All (ZIP)",
            data=zip_buffer,
            file_name="checked_files.zip",
            mime="application/zip"
        )
